import string

import psycopg2 as lib
import openpyxl
import numpy as np

def read_column():
    wookbook = openpyxl.load_workbook("Patients.xlsx")
    worksheet = wookbook.active

    j = 0
    max_row = 60
    full_data = []
    data = []
    for j in range(1,worksheet.max_column - 1):
        for i in range(2, max_row+1):
            data.append(worksheet.cell(row=i, column=j).value)

        full_data.append(data)
        data = []

    return full_data

def insert_patients():
    # db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
    # mycursor1 = db1.cursor()
    data = []
    data = read_column()
    for i in range(len(data[0])):
        ID_patients = data[0][i]
        gender = data[5][i]
        data_born = data[6][i]
        db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
        mycursor1 = db1.cursor()
        try:
            # mycursor1.execute('SELECT *  FROM insert_new_patients(%(p1)s,%(p2)s,%(p3)s,%(p4)s);', {"p1":first_name,"p2": last_name, "p3": doctor_id,"p4": data_born})

            mycursor1.execute('INSERT INTO patients values(%(p1)s,%(p2)s,%(p3)s);',{"p1":ID_patients,"p2": gender, "p3": data_born})

            db1.commit()
        except (Exception, lib.DatabaseError) as error:
            print(error)
        finally:
            mycursor1.close()
            if db1 is not None:
                db1.close()

def insert_disease_covid19():
    data = []
    data = read_column()
    for i in range(len(data[0])):
        ID_patients = data[0][i]
        data_order = str(data[15][i])
        confirmation = data[16][i]
        KT = data[17][i]

        db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
        mycursor1 = db1.cursor()
        try:

            mycursor1.execute('INSERT INTO disease_covid19 values(%(p1)s,%(p2)s,%(p3)s,%(p4)s,%(p5)s);',{"p1":ID_patients,"p2":ID_patients,"p3": data_order[0:10], "p4": confirmation,"p5": KT})

            db1.commit()
        except (Exception, lib.DatabaseError) as error:
            print(error)
        finally:
            mycursor1.close()
            if db1 is not None:
                db1.close()

def insert_pulmonary_auscultation():

    data = read_column()
    for i in range(len(data[0])):
        ID_patients = data[0][i]
        ch1 = data[27][i]
        ch2 = data[28][i]
        ch3 = data[29][i]
        ch4 = data[30][i]
        ch5 = data[31][i]
        ch6 = data[32][i]
        comments = str(data[33][i])

        db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
        mycursor1 = db1.cursor()
        try:

            mycursor1.execute('INSERT INTO pulmonary_auscultation values(%(p1)s,%(p2)s,%(p3)s,%(p4)s,%(p5)s,%(p6)s,%(p7)s,%(p8)s,%(p9)s);',{"p1":ID_patients,"p2":ID_patients,"p3": ch1, "p4": ch2,"p5": ch3,"p6":ch4,"p7":ch5,"p8": ch6, "p9": comments})

            db1.commit()
        except (Exception, lib.DatabaseError) as error:
            print(error)
        finally:
            mycursor1.close()
            if db1 is not None:
                db1.close()

def insert_vaccine_covid19():

    data = read_column()
    for i in range(len(data[0])):
        ID_patients = data[0][i]
        data_vaccine = str(data[20][i])
        vaccine = data[19][i]


        db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
        mycursor1 = db1.cursor()
        try:

            mycursor1.execute('INSERT INTO vaccine_covid19 values(%(p1)s,%(p2)s,%(p3)s,%(p4)s);',{"p1":ID_patients,"p2":ID_patients,"p3": data_vaccine, "p4": vaccine})

            db1.commit()
        except (Exception, lib.DatabaseError) as error:
            print(error)
        finally:
            mycursor1.close()
            if db1 is not None:
                db1.close()


def insert_orders():

    data = read_column()
    for i in range(len(data[0])):
        ID_patients = data[0][i]
        data_order = str(data[1][i])
        purpose = data[12][i]
        doctor = data[4][i]
        place = data[3][i]

        db1 = lib.connect(dbname="lungs", user="postgres", password="150120", host="127.0.0.1", port="5432")
        mycursor1 = db1.cursor()
        try:

            mycursor1.execute('INSERT INTO orders values(%(p1)s,%(p2)s,%(p3)s,%(p4)s,%(p5)s,%(p6)s);',{"p1":ID_patients,"p2":ID_patients,"p3": data_order[0:10], "p4": purpose,"p5": doctor, "p6": place})

            db1.commit()
        except (Exception, lib.DatabaseError) as error:
            print(error)
        finally:
            mycursor1.close()
            if db1 is not None:
                db1.close()

def len_txt(path):
    path = str(path) + '.txt'
    with open(path, 'r',encoding = "ISO-8859-1") as fp:
      x = len(fp.readlines())
      print('Total lines:', x)
      fp.close()
    return x


def upload_txt(path, x, column):
    path = str(path) + '.txt'
    print("Файл", path)
    with open(path, "r", encoding="ISO-8859-1") as file:
        for i in range(16):
            line = file.readline()

        x -= 17

        line = file.readline()

        data_txt = np.zeros(x)
        line = line.split()
        data_txt[0] = line[column]
        for i in range(1, x):
            line = file.readline()
            line = line.split()
            # print(line[0])
            data_txt[i] = (float(line[column]))
        file.close()
    return data_txt





if __name__ == '__main__':
    # data = []
    # data = read_column()
    # print(data)
    insert_patients()
    insert_disease_covid19()
    insert_pulmonary_auscultation()
    insert_vaccine_covid19()
    insert_orders()
    file_number = 1
    channel = 1
    column_amount = len_txt(file_number)
    signal = upload_txt(file_number, column_amount, channel)
    print(signal)



